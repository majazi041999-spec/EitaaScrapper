#!/usr/bin/env python3
"""
Eitaa multi-thread scraper (Excel input, link range)
Strategy: Direct Sequential Access + Smart Retry
"""

from __future__ import annotations

import concurrent.futures
import dataclasses
import logging
import re
import subprocess
import sys
import threading
import time
from pathlib import Path
from typing import Iterable, Optional

from openpyxl import load_workbook
from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
from playwright.sync_api import sync_playwright

BASE_URL = "https://eitaa.com"
DEFAULT_TIMEOUT_MS = 35000
REQUEST_WAIT_SECONDS = 1.25  # between direct post fetches
POST_RETRY_COUNT = 2

DEFAULT_THREADS = 8
DEFAULT_OUTPUT = Path("result_fa.txt").resolve()
DEFAULT_LOG_DIR = Path("logs").resolve()

LOGGER_NAME = "eitaa_scraper"
_DIGIT_MAP = str.maketrans("۰۱۲۳۴۵۶۷۸۹", "0123456789")


@dataclasses.dataclass
class ChannelTask:
    channel: str
    start_link: str
    end_link: str
    start_id: int
    end_id: int


@dataclasses.dataclass
class ChannelResult:
    channel: str
    start_link: str
    end_link: str
    status: str = "موفق"
    error_message: str = ""

    total_posts: int = 0
    total_views: int = 0

    non_forward_posts: int = 0
    non_forward_views: int = 0

    forwarded_posts: int = 0
    forwarded_views: int = 0

    photo_posts: int = 0
    video_posts: int = 0
    sticker_posts: int = 0

    not_found_posts: int = 0
    service_posts: int = 0
    failed_posts: int = 0


class MaxLevelFilter(logging.Filter):
    def __init__(self, max_level: int) -> None:
        super().__init__()
        self.max_level = max_level

    def filter(self, record: logging.LogRecord) -> bool:
        return record.levelno <= self.max_level


def setup_logger(log_dir: Path) -> logging.Logger:
    log_dir.mkdir(parents=True, exist_ok=True)
    logger = logging.getLogger(LOGGER_NAME)
    logger.setLevel(logging.DEBUG)
    logger.handlers.clear()

    formatter = logging.Formatter(
        "%(asctime)s | %(levelname)-8s | %(threadName)s | %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    verbose_handler = logging.FileHandler(log_dir / "verbose.log", encoding="utf-8")
    verbose_handler.setLevel(logging.DEBUG)
    verbose_handler.addFilter(MaxLevelFilter(logging.INFO))
    verbose_handler.setFormatter(formatter)

    warning_handler = logging.FileHandler(log_dir / "warning.log", encoding="utf-8")
    warning_handler.setLevel(logging.WARNING)
    warning_handler.addFilter(MaxLevelFilter(logging.WARNING))
    warning_handler.setFormatter(formatter)

    error_handler = logging.FileHandler(log_dir / "error.log", encoding="utf-8")
    error_handler.setLevel(logging.ERROR)
    error_handler.setFormatter(formatter)

    stream_handler = logging.StreamHandler(sys.stdout)
    stream_handler.setLevel(logging.INFO)
    stream_handler.setFormatter(formatter)

    logger.addHandler(verbose_handler)
    logger.addHandler(warning_handler)
    logger.addHandler(error_handler)
    logger.addHandler(stream_handler)
    return logger


def _sanitize_path_input(raw: str) -> Path:
    text = raw.strip().strip('"').strip("'")
    return Path(text).expanduser().resolve()


def _parse_channel_post_link(link: str) -> tuple[str, int]:
    value = link.strip()
    pattern = r"https?://(?:www\.)?eitaa\.com/(?:s/)?([A-Za-z0-9_]+)/([0-9]+)"
    m = re.fullmatch(pattern, value)
    if not m:
        raise ValueError(f"Invalid eitaa post link: {link}")
    return f"@{m.group(1)}", int(m.group(2))


def read_tasks_from_excel(excel_path: Path) -> list[ChannelTask]:
    wb = load_workbook(filename=excel_path, read_only=True, data_only=True)
    ws = wb.active
    tasks: list[ChannelTask] = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
        if not row:
            continue

        c1 = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ""
        c2 = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        c3 = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""

        if not c1 and not c2 and not c3:
            continue

        if row_idx == 1 and c1.lower() in {"channel", "کانال"}:
            continue

        if not (c1 and c2 and c3):
            raise ValueError(f"Row {row_idx}: all 3 columns are required (channel, end link, start link).")

        # user contract: column2=end link, column3=start link
        end_channel, end_id = _parse_channel_post_link(c2)
        start_channel, start_id = _parse_channel_post_link(c3)

        channel = c1 if c1.startswith("@") else f"@{c1}"
        if start_channel.lower() != channel.lower() or end_channel.lower() != channel.lower():
            raise ValueError(
                f"Row {row_idx}: channel mismatch. column1={channel}, end={end_channel}, start={start_channel}"
            )

        if start_id > end_id:
            start_id, end_id = end_id, start_id
            c3, c2 = c2, c3

        tasks.append(ChannelTask(channel=channel, start_link=c3, end_link=c2, start_id=start_id, end_id=end_id))

    wb.close()
    return tasks


def _normalize_number_text(text: str) -> str:
    return (
        text.translate(_DIGIT_MAP)
        .replace(",", "")
        .replace(" ", "")
        .replace("٬", "")
        .replace("٫", ".")
        .lower()
    )


def _parse_views(text: str) -> Optional[int]:
    if not text:
        return None

    cleaned = _normalize_number_text(text)
    match = re.search(r"(\d+(?:\.\d+)?)([km]?)", cleaned)
    if not match:
        return None

    number = float(match.group(1))
    suffix = match.group(2)
    if suffix == "k":
        number *= 1000
    elif suffix == "m":
        number *= 1_000_000

    return int(number)


def _extract_single_post_payload(page, post_id: int) -> dict:
    script = """
    ({ targetPostId }) => {
      const wraps = Array.from(document.querySelectorAll('.tgme_widget_message_wrap, .etme_widget_message_wrap, .widget_message_wrap'));

      let target = null;
      for (const wrap of wraps) {
        const msg = wrap.querySelector('[data-post], .tgme_widget_message, .etme_widget_message') || wrap;
        const postKey = msg.getAttribute('data-post') || wrap.getAttribute('data-post') || msg.id || wrap.id || '';
        const idMatch = /\/(\d+)$/.exec(postKey);
        const pid = idMatch ? Number(idMatch[1]) : null;
        if (pid === targetPostId) {
          target = wrap;
          break;
        }
      }

      if (!target) {
        return { status: 'not_found' };
      }

      if (target.querySelector('.etme_widget_message_service, .tgme_widget_message_service, .service_message')) {
        return { status: 'service_message' };
      }

      const viewsNode = target.querySelector('.tgme_widget_message_views, .etme_widget_message_views, [class*="message_views"]');
      const viewsDataCount = viewsNode ? (viewsNode.getAttribute('data-count') || '') : '';
      const viewsText = viewsNode ? (viewsNode.textContent || '') : '';

      const isForwarded = !!target.querySelector('.tgme_widget_message_forwarded_from, .etme_widget_message_forwarded_from, [class*="forwarded"]');
      const hasPhoto = !!target.querySelector('a.tgme_widget_message_photo_wrap, .tgme_widget_message_photo_wrap, .etme_widget_message_photo_wrap, [class*="photo_wrap"], img');
      const hasVideo = !!target.querySelector('video, .tgme_widget_message_video, .etme_widget_message_video, [class*="video"]');
      const hasSticker = !!target.querySelector('.tgme_widget_message_sticker, .etme_widget_message_sticker, [class*="sticker"]');

      return {
        status: 'ok',
        post_id: targetPostId,
        views_data_count: viewsDataCount,
        views_text: viewsText,
        is_forwarded: isForwarded,
        has_photo: hasPhoto,
        has_video: hasVideo,
        has_sticker: hasSticker
      };
    }
    """
    return page.evaluate(script, {"targetPostId": post_id})


def _navigate_to_channel_home(page, channel: str, logger: logging.Logger) -> None:
    # warmup and cookies
    url = f"{BASE_URL}/s/{channel.lstrip('@')}"
    logger.info("Opening channel warmup %s", url)
    page.goto(url, wait_until="domcontentloaded", timeout=DEFAULT_TIMEOUT_MS)
    page.wait_for_timeout(1800)


def _fetch_post(page, channel: str, post_id: int) -> dict:
    post_url = f"{BASE_URL}/{channel.lstrip('@')}/{post_id}"
    page.goto(post_url, wait_until="domcontentloaded", timeout=DEFAULT_TIMEOUT_MS)
    page.wait_for_timeout(600)
    return _extract_single_post_payload(page, post_id)


def _apply_post_to_result(result: ChannelResult, payload: dict, logger: logging.Logger) -> None:
    if payload.get("status") != "ok":
        return

    views: Optional[int] = None
    data_count = str(payload.get("views_data_count") or "").strip()
    if data_count.isdigit():
        views = int(data_count)
    else:
        views = _parse_views(str(payload.get("views_text") or ""))

    if views is None:
        logger.warning("[%s] post=%s skipped (views unreadable: '%s')", result.channel, payload.get("post_id"), payload.get("views_text"))
        result.failed_posts += 1
        return

    result.total_posts += 1
    result.total_views += views

    if bool(payload.get("is_forwarded")):
        result.forwarded_posts += 1
        result.forwarded_views += views
    else:
        result.non_forward_posts += 1
        result.non_forward_views += views

    if bool(payload.get("has_photo")):
        result.photo_posts += 1
    if bool(payload.get("has_video")):
        result.video_posts += 1
    if bool(payload.get("has_sticker")):
        result.sticker_posts += 1


def _scan_channel_posts_sequential(page, task: ChannelTask, logger: logging.Logger) -> ChannelResult:
    result = ChannelResult(channel=task.channel, start_link=task.start_link, end_link=task.end_link)
    errors: set[int] = set()

    total_targets = task.end_id - task.start_id + 1
    logger.info("[%s] Sequential scan range: %d -> %d (%d posts)", task.channel, task.start_id, task.end_id, total_targets)

    for post_id in range(task.start_id, task.end_id + 1):
        try:
            payload = _fetch_post(page, task.channel, post_id)
        except Exception:
            errors.add(post_id)
            logger.warning("[%s] post=%d failed in first pass", task.channel, post_id)
            time.sleep(REQUEST_WAIT_SECONDS)
            continue

        status = payload.get("status")
        if status == "not_found":
            result.not_found_posts += 1
        elif status == "service_message":
            result.service_posts += 1
        elif status == "ok":
            _apply_post_to_result(result, payload, logger)
        else:
            errors.add(post_id)

        time.sleep(REQUEST_WAIT_SECONDS)

    if errors:
        logger.info("[%s] Retry pass for %d failed posts", task.channel, len(errors))

    for _ in range(POST_RETRY_COUNT):
        if not errors:
            break
        current = sorted(errors)
        errors.clear()
        for post_id in current:
            try:
                payload = _fetch_post(page, task.channel, post_id)
                status = payload.get("status")
                if status == "not_found":
                    result.not_found_posts += 1
                elif status == "service_message":
                    result.service_posts += 1
                elif status == "ok":
                    _apply_post_to_result(result, payload, logger)
                else:
                    errors.add(post_id)
            except Exception:
                errors.add(post_id)
            time.sleep(max(0.6, REQUEST_WAIT_SECONDS / 2))

    result.failed_posts += len(errors)

    logger.info(
        "[%s] Finished: kept_posts=%d total_views=%d | not_found=%d service=%d failed=%d",
        task.channel,
        result.total_posts,
        result.total_views,
        result.not_found_posts,
        result.service_posts,
        result.failed_posts,
    )
    return result


def _ensure_playwright_browser(logger: logging.Logger) -> None:
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            browser.close()
        return
    except Exception:
        logger.warning("Chromium not found. Trying to install Playwright Chromium automatically...")
        candidates = [
            [sys.executable, "-m", "playwright", "install", "chromium"],
            ["playwright", "install", "chromium"],
        ]
        last_error: Optional[Exception] = None
        for cmd in candidates:
            try:
                subprocess.run(cmd, check=True)
                return
            except Exception as exc:
                last_error = exc
        raise RuntimeError(f"could not install chromium automatically: {last_error}")


def process_channel(task: ChannelTask, headless: bool, logger: logging.Logger) -> ChannelResult:
    logger.info("[%s] Worker started in %s", task.channel, threading.current_thread().name)

    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=headless)
            context = browser.new_context(viewport={"width": 1366, "height": 920})
            page = context.new_page()
            page.set_default_timeout(DEFAULT_TIMEOUT_MS)

            _navigate_to_channel_home(page, task.channel, logger)
            result = _scan_channel_posts_sequential(page, task, logger)

            context.close()
            browser.close()
            return result

    except PlaywrightTimeoutError as exc:
        logger.error("[%s] Timeout error: %s", task.channel, exc)
        return ChannelResult(task.channel, task.start_link, task.end_link, status="خطا", error_message=f"Timeout: {exc}")
    except Exception as exc:
        logger.exception("[%s] Unhandled error", task.channel)
        return ChannelResult(task.channel, task.start_link, task.end_link, status="خطا", error_message=str(exc))


def write_report(output_path: Path, results: Iterable[ChannelResult], elapsed_seconds: float) -> None:
    items = sorted(results, key=lambda x: x.channel.lower())

    total_channels = len(items)
    success_count = sum(1 for x in items if x.status == "موفق")
    fail_count = total_channels - success_count

    grand_posts = sum(x.total_posts for x in items)
    grand_views = sum(x.total_views for x in items)
    grand_nonf_posts = sum(x.non_forward_posts for x in items)
    grand_nonf_views = sum(x.non_forward_views for x in items)
    grand_fwd_posts = sum(x.forwarded_posts for x in items)
    grand_fwd_views = sum(x.forwarded_views for x in items)
    grand_photos = sum(x.photo_posts for x in items)
    grand_videos = sum(x.video_posts for x in items)
    grand_stickers = sum(x.sticker_posts for x in items)
    grand_not_found = sum(x.not_found_posts for x in items)
    grand_service = sum(x.service_posts for x in items)
    grand_failed = sum(x.failed_posts for x in items)

    lines: list[str] = []
    lines.append("=" * 84)
    lines.append("گزارش نهایی اسکرپر ایتا (Direct Sequential + Retry)")
    lines.append("=" * 84)
    lines.append(f"زمان تولید گزارش: {time.strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"تعداد کانال‌ها: {total_channels}")
    lines.append(f"موفق: {success_count} | خطادار: {fail_count}")
    lines.append(f"کل پست‌های محاسبه‌شده: {grand_posts:,} | کل ویو: {grand_views:,}")
    lines.append(f"پست غیرفوروارد: {grand_nonf_posts:,} | ویو غیرفوروارد: {grand_nonf_views:,}")
    lines.append(f"پست فوروارد: {grand_fwd_posts:,} | ویو فوروارد: {grand_fwd_views:,}")
    lines.append(f"رسانه -> عکس: {grand_photos:,} | ویدیو: {grand_videos:,} | استیکر: {grand_stickers:,}")
    lines.append(f"پست موجود نبود: {grand_not_found:,} | پیام سرویس: {grand_service:,} | خطای نهایی: {grand_failed:,}")
    lines.append(f"زمان اجرا: {elapsed_seconds:.2f} ثانیه")
    lines.append("")

    for idx, item in enumerate(items, start=1):
        lines.append("-" * 84)
        lines.append(f"کانال #{idx}: {item.channel}")
        lines.append(f"لینک شروع: {item.start_link}")
        lines.append(f"لینک پایان: {item.end_link}")
        lines.append(f"وضعیت: {item.status}")
        if item.status == "موفق":
            lines.append(f"کل پست‌های محاسبه‌شده: {item.total_posts:,}")
            lines.append(f"کل ویو: {item.total_views:,}")
            lines.append(f"غیرفوروارد -> پست: {item.non_forward_posts:,} | ویو: {item.non_forward_views:,}")
            lines.append(f"فوروارد -> پست: {item.forwarded_posts:,} | ویو: {item.forwarded_views:,}")
            lines.append(f"رسانه -> عکس: {item.photo_posts:,} | ویدیو: {item.video_posts:,} | استیکر: {item.sticker_posts:,}")
            lines.append(f"حذف‌شده/ناموجود -> not_found: {item.not_found_posts:,} | service: {item.service_posts:,} | failed: {item.failed_posts:,}")
        else:
            lines.append(f"جزئیات خطا: {item.error_message}")
        lines.append("")

    lines.append("=" * 84)
    lines.append("پایان گزارش")
    output_path.write_text("\n".join(lines), encoding="utf-8")


def _ask_user_inputs() -> tuple[Path, int, bool]:
    print("\n=== Eitaa Scraper (Excel Mode) ===")
    print("Excel format (3 columns per row):")
    print("1) Channel    2) END post link    3) START post link")
    print("Example link: https://eitaa.com/mychannel/12345")
    print("Tip: you can drag & drop the Excel file path into terminal.\n")

    excel_path = _sanitize_path_input(input("Enter Excel file path: "))
    threads_raw = input(f"Enter thread count [default {DEFAULT_THREADS}]: ").strip()
    headful_raw = input("Run browser visible? (y/N): ").strip().lower()

    threads = int(threads_raw) if threads_raw else DEFAULT_THREADS
    headful = headful_raw in {"y", "yes", "1"}
    return excel_path, max(1, threads), headful


def main() -> int:
    excel_path, threads, headful = _ask_user_inputs()

    logger = setup_logger(DEFAULT_LOG_DIR)
    if not excel_path.exists():
        logger.error("Excel file not found: %s", excel_path)
        return 1

    try:
        tasks = read_tasks_from_excel(excel_path)
    except Exception as exc:
        logger.error("Cannot read Excel tasks: %s", exc)
        return 1

    if not tasks:
        logger.error("No valid rows found in excel file: %s", excel_path)
        return 1

    try:
        _ensure_playwright_browser(logger)
    except Exception as exc:
        logger.error("Failed to ensure Playwright Chromium: %s", exc)
        logger.error("For exe mode, run once on target machine: playwright install chromium")
        return 1

    logger.info("Loaded %d channel rows from %s", len(tasks), excel_path)
    logger.info("Thread count: %d", threads)
    logger.info("Result file: %s", DEFAULT_OUTPUT)
    logger.info("Log directory: %s", DEFAULT_LOG_DIR)
    logger.info("Mode: Direct Sequential Access + Retry | per-post wait=%.2fs", REQUEST_WAIT_SECONDS)

    start_ts = time.time()
    results: list[ChannelResult] = []

    with concurrent.futures.ThreadPoolExecutor(max_workers=threads) as executor:
        future_map = {executor.submit(process_channel, task, not headful, logger): task.channel for task in tasks}
        for future in concurrent.futures.as_completed(future_map):
            channel = future_map[future]
            try:
                item = future.result()
            except Exception as exc:
                logger.exception("[%s] unexpected worker failure", channel)
                item = ChannelResult(channel, "-", "-", status="خطا", error_message=str(exc))
            results.append(item)

    elapsed = time.time() - start_ts
    write_report(DEFAULT_OUTPUT, results, elapsed)
    logger.info("Report written to %s", DEFAULT_OUTPUT)
    logger.info("Done in %.2f seconds", elapsed)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
